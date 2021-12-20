import torch
import torch.nn as nn
from pytorch_lightning.core.lightning import LightningModule
import matplotlib.pyplot as plt
from os.path import join as opj
import sys
import os
from torchmetrics import F1, Accuracy, MeanSquaredError, MeanAbsoluteError
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

import config
from UNet3d import UNet3d


class ExperimentModule(LightningModule):

    def __init__(self, learning_mode, pretrained, distortions):
        super(ExperimentModule, self).__init__()

        self.unet = UNet3d()

        if pretrained:
            if distortions:
                if sys.argv[3] == "light":
                    self.unet.load_state_dict(torch.load(opj(config.dirpath, "distlightpretrained_model.pt")))
                if sys.argv[3] == "normal":
                    self.unet.load_state_dict(torch.load(opj(config.dirpath, "distnormalpretrained_model.pt")))
                if sys.argv[3] == "strong":
                    self.unet.load_state_dict(torch.load(opj(config.dirpath, "diststrongpretrained_model.pt")))
                if sys.argv[3] == "dist":
                    self.unet.load_state_dict(torch.load(opj(config.dirpath, "distallpretrained_model.pt")))
            else:
                self.unet.load_state_dict(torch.load(opj(config.dirpath, "nodistpretrained_model.pt")))

        self.learning_mode = learning_mode

        if self.learning_mode == "n_peaks":
            self.out_block = nn.Conv3d(config.num_filter, config.out_dim_peaks, kernel_size=1)
            self.loss = nn.L1Loss()
            self.metric = F1(threshold=0.5)
            self.metric2 = Accuracy()
        elif self.learning_mode == "regression":
            self.out_block = nn.Conv3d(config.num_filter, config.out_dim_regression, kernel_size=1)
            self.loss = nn.MSELoss()
            self.metric = MeanSquaredError()
            self.metric2 = MeanAbsoluteError()
        elif self.learning_mode == "segmentation":
            self.out_block = nn.Conv3d(config.num_filter, config.out_dim_segmentation, kernel_size=1)
            self.loss = nn.L1Loss()
            self.metric = F1(threshold=0.5)
            self.metric2 = Accuracy()
        else:
            raise Exception("unknown learning mode")

    def forward(self, z):
        y = self.unet(z)
        return self.out_block(y)

    def training_step(self, batch, batch_idx):
        target = batch['target']
        input = batch['input']

        output = self.forward(input)

        train_loss = self.loss(output, target)
        self.log('Loss/Train', train_loss)
        return train_loss

    def validation_step(self, batch, batch_idx):
        target = batch['target']
        input = batch['input']

        output = self.forward(input)

        if 0:
            axial_middle = output.shape[2] // 2
            plt.figure('Showing the datasets')

            plt.subplot(2, 2, 1).set_axis_off()
            plt.title("Input")
            plt.imshow(input.cpu().detach().numpy()[0, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 2).set_axis_off()
            plt.title("Output 0.5")
            plt.imshow((output.cpu().detach().numpy() > 0.5).astype(int)[0, 1, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 3).set_axis_off()
            plt.title("Target")
            plt.imshow(target.cpu().detach().numpy()[0, 1, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 4).set_axis_off()
            plt.title("Output")
            plt.imshow(output.cpu().detach().numpy()[0, 1, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.show()

        val_loss = self.loss(output, target)
        self.log('val_loss', val_loss)

        print("\n")
        print("Validation Loss:", val_loss)
        print("\n")

        return val_loss

    def validation_epoch_end(self, validation_step_outputs):
        pass

    def test_step(self, batch, batch_idx):
        target = batch['target']
        input = batch['input']
        log = np.zeros([len(target[:,0,0,0,0]),4])

        for i in range(len(target[:,0,0,0,0])):

            output = self.forward(torch.unsqueeze(input[i, :, :, :, :], 0))

            test_loss = self.loss(target[i, :, :, :, :], output[0, :, :, :, :])
            self.log('test_loss', test_loss, on_step=True)

            metric = self.metric(output[0, :, :, :, :].cpu(), target[i, :, :, :, :].int().cpu())
            self.log('f1ORMSE', metric, on_step=True)

            metric2 = self.metric2(output[0, :, :, :, :].cpu(), target[i, :, :, :, :].int().cpu())
            self.log('AccuracyORMAE', metric2, on_step=True)

            log[i,0] = batch_idx
            log[i,1] = test_loss.item()
            log[i,2] = metric.item()
            log[i,3] = metric2.item()

        df = pd.DataFrame(log)

        if not os.path.exists(config.log_path):
            wb = openpyxl.Workbook()
            wb.title = "RawData"
            wb.save(filename=config.log_path)
        writer = pd.ExcelWriter(config.log_path, engine='openpyxl')
        writer.book = load_workbook(config.log_path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(config.log_path)
        df.to_excel(writer, sheet_name="RawData", index=True, header=False, startrow=len(reader) + 1)
        writer.close()

    def on_test_epoch_end(self):
        pass

    def configure_optimizers(self):

        lr = config.lr
        optimizer = torch.optim.Adam(filter(lambda p: p.requires_grad, self.parameters()), lr=lr)

        return {'optimizer': optimizer}

