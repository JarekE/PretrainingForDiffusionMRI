# This CNN is build on the U-Net architecture
import torch
import torch.nn as nn
from torch.nn import functional as F
from pytorch_lightning.core.lightning import LightningModule
from pytorch_lightning.metrics.classification import f_beta
import math
from pytorch_lightning.metrics import Accuracy
import matplotlib.pyplot as plt
import sys
import numpy as np
from os.path import join as opj
from pandas import DataFrame
from openpyxl import load_workbook
import pandas as pd

import dMRIconfig


def conv_block(in_dim, out_dim, act_fn):
    model = nn.Sequential(
        nn.Conv3d(in_dim, out_dim, kernel_size=3, stride=1, padding=1),
        nn.InstanceNorm3d(out_dim),
        act_fn,
    )
    return model


def up_conv(in_dim, out_dim, act_fn):
    model = nn.Sequential(
        nn.ConvTranspose3d(in_dim,out_dim, kernel_size=3, stride=2, padding=1,output_padding=1),
        nn.InstanceNorm3d(out_dim),
        act_fn,
    )
    return model


def double_conv_block(in_dim, out_dim, act_fn):
    model = nn.Sequential(
        conv_block(in_dim, out_dim, act_fn),
        conv_block(out_dim, out_dim, act_fn),
    )
    return model


def out_block(in_dim,out_dim):
    model = nn.Sequential(
        nn.Conv3d(in_dim, out_dim, kernel_size=1, stride=1, padding=0),
    )
    return model


class Unet3d(LightningModule):

    def __init__(self, in_dim, out_dim, num_filter, out_dim_pretraining, out_dim_regression,
                 out_dim_classification, learning_modus="out", pretraining="nopre"):
        super(Unet3d, self).__init__()
        self.accuracy = Accuracy()
        self.MSE = nn.MSELoss()
        self.f_beta = f_beta.Fbeta
        self.in_dim = in_dim
        self.out_dim = out_dim
        self.num_filter = num_filter
        self.learning_modus = learning_modus
        self.pretraining = pretraining
        act_fn = nn.ReLU(inplace=True)

        self.down_1 = double_conv_block(self.in_dim, self.num_filter, act_fn)
        self.pool_1 = nn.MaxPool3d(kernel_size=2, stride=2, padding=0)
        self.down_2 = double_conv_block(self.num_filter, self.num_filter * 2, act_fn)
        self.pool_2 = nn.MaxPool3d(kernel_size=2, stride=2, padding=0)
        self.down_3 = double_conv_block(self.num_filter * 2, self.num_filter * 4, act_fn)
        self.pool_3 = nn.MaxPool3d(kernel_size=2, stride=2, padding=0)

        self.bridge = double_conv_block(self.num_filter * 4, self.num_filter * 8, act_fn)

        self.trans_1 = up_conv(self.num_filter * 8, self.num_filter * 8, act_fn)
        self.up_1 = double_conv_block(self.num_filter * 12, self.num_filter * 4, act_fn)

        self.trans_2 = up_conv(self.num_filter * 4, self.num_filter * 4, act_fn)
        self.up_2 = double_conv_block(self.num_filter * 6, self.num_filter * 2, act_fn)

        self.trans_3 = up_conv(self.num_filter * 2, self.num_filter * 2, act_fn)
        self.up_3 = double_conv_block(self.num_filter * 3, self.num_filter, act_fn)

        self.out_pretraining = out_block(self.num_filter, out_dim_pretraining)
        self.out = out_block(self.num_filter, out_dim)
        self.out_regression = out_block(self.num_filter, out_dim_regression)
        self.out_classification = out_block(self.num_filter, out_dim_classification)

        self.epochs_list = []
        self.f1_list = []
        self.acc_list = []
        self.loss_list = []
        self.learning_modus_list = []
        self.pretraining_list = []

        self.epochs_list_test = []
        self.f1_list_test = []
        self.acc_list_test = []
        self.loss_list_test = []
        self.learning_modus_list_test = []
        self.pretraining_list_test = []


    def forward(self, x):

        down_1 = self.down_1(x)
        pool_1 = self.pool_1(down_1)
        down_2 = self.down_2(pool_1)
        pool_2 = self.pool_2(down_2)
        down_3 = self.down_3(pool_2)
        pool_3 = self.pool_3(down_3)

        bridge = self.bridge(pool_3)

        trans_1 = self.trans_1(bridge)
        concat_1 = torch.cat([trans_1, down_3], dim=1)
        up_1 = self.up_1(concat_1)
        trans_2 = self.trans_2(up_1)
        concat_2 = torch.cat([trans_2, down_2], dim=1)
        up_2 = self.up_2(concat_2)
        trans_3 = self.trans_3(up_2)
        concat_3 = torch.cat([trans_3, down_1], dim=1)
        up_3 = self.up_3(concat_3)

        if self.learning_modus == "out" or self.learning_modus == "out_number":
            out = self.out(up_3)
            return out
        elif self.learning_modus == "out_regression":
            out_regression = self.out_regression(up_3)
            return out_regression
        elif self.learning_modus == "out_classification":
            out_classification = self.out_classification(up_3)
            return out_classification
        else:
            raise Exception("unknown second argument")


    def training_step(self, batch, batch_idx):

        label = batch['label']
        input = batch['input']

        label = label.cuda(label.device.index)
        input = input.cuda(input.device.index)

        output = self.forward(input)

        if 0:
            axial_middle = output.shape[2] // 2
            plt.figure('Showing the datasets')

            plt.subplot(2, 2, 1).set_axis_off()
            plt.title("Input")
            plt.imshow(input.cpu().detach().numpy()[0, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 2).set_axis_off()
            plt.title("Input")
            plt.imshow(input.cpu().detach().numpy()[2, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 3).set_axis_off()
            plt.title("Output")
            plt.imshow(output.cpu().detach().numpy()[0, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.subplot(2, 2, 4).set_axis_off()
            plt.title("Output")
            plt.imshow(output.cpu().detach().numpy()[2, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

            plt.show()


        if self.learning_modus == "out" or self.learning_modus == "out_number":
            train_loss_out = F.l1_loss(label, output)
            self.train_loss_out = train_loss_out
            return train_loss_out
        elif self.learning_modus == "out_regression":
            train_loss_reg = self.MSE(label, output)
            self.train_loss_reg = train_loss_reg
            return train_loss_reg
        elif self.learning_modus == "out_classification":
            train_loss_class = F.l1_loss(label, output)
            self.train_loss_class = train_loss_class
            return train_loss_class
        else:
            raise Exception("unknown second argument")


    def validation_step(self, batch, batch_idx):

        label = batch['label']
        input = batch['input']

        label = label.cuda(label.device.index)
        input = input.cuda(input.device.index)

        output = self.forward(input)

        if 1:
            if self.learning_modus == "out_number":
                axial_middle = output.shape[2] // 2
                plt.figure('Showing the datasets')

                plt.subplot(3, 1, 1).set_axis_off()
                plt.title("Input")
                plt.imshow(input.cpu().detach().numpy()[0, 10, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.subplot(3, 1, 2).set_axis_off()
                plt.title("Output")
                plt.imshow(output.cpu().detach().numpy()[0, 2, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.subplot(3, 1, 3).set_axis_off()
                plt.title("Label")
                plt.imshow(label.cpu().detach().numpy()[0, 2, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.show()
            elif self.learning_modus == "out":
                axial_middle = output.shape[2] // 2
                plt.figure('Showing the datasets')

                plt.subplot(3, 1, 1).set_axis_off()
                plt.title("Input")
                plt.imshow(input.cpu().detach().numpy()[0, 11, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.subplot(3, 1, 2).set_axis_off()
                plt.title("Output")
                plt.imshow((output.cpu().detach().numpy() > 0.5).astype(int)[0, 1, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.subplot(3, 1, 3).set_axis_off()
                plt.title("Label")
                plt.imshow(label.cpu().detach().numpy()[0, 1, :, axial_middle, :].T, cmap='gray', origin='lower')

                plt.show()
            else:
                print("This output will be here in the near future!")


        if self.learning_modus == "out" or self.learning_modus == "out_number":
            f1 = self.f_beta(num_classes=4, threshold=0.5, beta=1)
            f1_acc = f1(output.cpu(), label.cpu())
            acc = self.accuracy(output.cpu(), label.cpu())
            val_loss = F.l1_loss(label, output)
            self.log('val_loss', val_loss)

            return val_loss, f1_acc, acc

        if self.learning_modus == "out_classification":
            f1 = self.f_beta(num_classes=8, threshold=0.5, beta=1)
            f1_acc = f1(output.cpu(), label.cpu())
            acc = self.accuracy(output.cpu(), label.cpu())
            val_loss = F.l1_loss(label, output)
            self.log('val_loss', val_loss)

            return val_loss, f1_acc, acc

        if self.learning_modus == "out_regression":
            val_loss_reg = self.MSE(label, output)
            self.log('val_loss', val_loss_reg)

            return val_loss_reg, 0, 0


    def validation_epoch_end(self, validation_step_outputs):

        print("\nAktuelle Epoche:",self.current_epoch)
        print("Learning Modus:",self.learning_modus)
        print("Parameter: ", sys.argv[1], sys.argv[2])

        if self.learning_modus == "out_regression":
            print('Das Netzwerk hat auf den Validierungsdaten einen MSE loss von',
                  validation_step_outputs[0][0].item() * 100,
                  "%.")
        else:
            print('Das Netzwerk hat auf den Validierungsdaten einen f1 Score von', validation_step_outputs[0][1].item()*100,
                  "%.")

        #important to avoid empty cells in exel sheet
        if self.learning_modus == "out_regression":
            self.create_logger_file_training(0, 0, validation_step_outputs[0][0].item())
        else:
            if math.isnan(validation_step_outputs[0][1].item()) == True:
                self.create_logger_file_training(0, validation_step_outputs[0][2].item(), validation_step_outputs[0][0].item())
            else:
                self.create_logger_file_training(validation_step_outputs[0][1].item(), validation_step_outputs[0][2].item(),
                                        validation_step_outputs[0][0].item())

    def test_step(self, batch, batch_idx):

        label = batch['label']
        input = batch['input']

        label = label.cuda(label.device.index)
        input = input.cuda(input.device.index)

        for i in range(len(label[:,0,0,0,0])):

            output = self.forward(torch.unsqueeze(input[i,:,:,:,:], 0))

            if self.learning_modus == "out" or self.learning_modus == "out_number":
                f1 = self.f_beta(num_classes=4, threshold=0.5, beta=1)
                f1_acc = f1(output[0, :, :, :, :].cpu(), label[i, :, :, :, :].cpu())
                acc = self.accuracy(output[0, :, :, :, :].cpu(), label[i, :, :, :, :].cpu())
                val_loss = F.l1_loss(label[i, :, :, :, :], output[0, :, :, :, :])

            if self.learning_modus == "out_classification":
                f1 = self.f_beta(num_classes=8, threshold=0.5, beta=1)
                f1_acc = f1(output[0, :, :, :, :].cpu(), label[i, :, :, :, :].cpu())
                acc = self.accuracy(output[0, :, :, :, :].cpu(), label[i, :, :, :, :].cpu())
                val_loss = F.l1_loss(label[i, :, :, :, :], output[0, :, :, :, :])

            if self.learning_modus == "out_regression":
                val_loss = self.MSE(label[i, :, :, :, :], output[0, :, :, :, :])
                f1_acc = 0
                acc = 0

            #Save mask
            path = "/work/scratch/ecke/PretrainingForDiffusionMRI/Results/Endmasks"
            savepath_mask = opj(path, str(i)+str(batch_idx)+str(sys.argv[1])+str(sys.argv[2]))
            np.save(savepath_mask, output.cpu().numpy())

            self.create_logger_file_test(f1=f1_acc.item(), acc=acc.item(), loss=val_loss.item())

    def on_test_epoch_end(self) -> None:

        # The training data is currently not logged, due to the reason that the lists are getting deleted with the test phase
        # Maybe I will change this later
        # The test data is logged completly!
        '''
        save_path_f1 = '/work/scratch/ecke/PretrainingForDiffusionMRI/Results/Metrics_Table' + '/self_logs.xlsx'

        f1_logger = DataFrame({'Epoche': self.epochs_list,
                               'f1 Score': self.f1_list,
                               'Accuracy': self.acc_list,
                               'Loss': self.loss_list,
                               'Learning_modus': self.learning_modus_list,
                               'Pretraining': self.pretraining_list})

        writer = pd.ExcelWriter(save_path_f1, engine='openpyxl')
        writer.book = load_workbook(save_path_f1)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(save_path_f1)
        f1_logger.to_excel(writer, sheet_name="logs", index=True, header=False, startrow=len(reader) + 1)
        writer.close()
        '''
        save_path_f1 = '/work/scratch/ecke/PretrainingForDiffusionMRI/Results/Metrics_Table' + '/end_logs.xlsx'

        f1_logger = DataFrame({'Epoche': self.epochs_list_test,
                               'f1 Score': self.f1_list_test,
                               'Accuracy': self.acc_list_test,
                               'Loss': self.loss_list_test,
                               'Learning_modus': self.learning_modus_list_test,
                               'Pretraining': self.pretraining_list_test})

        writer = pd.ExcelWriter(save_path_f1, engine='openpyxl')
        writer.book = load_workbook(save_path_f1)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(save_path_f1)
        f1_logger.to_excel(writer, sheet_name="logs", index=True, header=False, startrow=len(reader) + 1)
        writer.close()


    def configure_optimizers(self):

        lr = dMRIconfig.lr
        optimizer = torch.optim.Adam(filter(lambda p: p.requires_grad, self.parameters()), lr=lr)

        return {'optimizer': optimizer}


    def create_logger_file_training(self, f1=0, acc=0, loss=0):

        # Lists are getting deletet with new instance of U-Net in the test phase.
        self.epochs_list.append(self.current_epoch)
        self.f1_list.append(f1)
        self.acc_list.append([acc])
        self.loss_list.append([loss])
        self.learning_modus_list.append(sys.argv[1])
        self.pretraining_list.append(sys.argv[2])


    def create_logger_file_test(self, f1=0, acc=0, loss=0):

        self.epochs_list_test.append(self.current_epoch)
        self.f1_list_test.append(f1)
        self.acc_list_test.append([acc])
        self.loss_list_test.append([loss])
        self.learning_modus_list_test.append(sys.argv[1])
        self.pretraining_list_test.append(sys.argv[2])
